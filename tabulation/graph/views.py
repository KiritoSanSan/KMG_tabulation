import calendar
import json
import logging
logger = logging.getLogger(__name__)
import re
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Count
from collections import defaultdict
# from django.urls import NoReverseMatch
from django.http import JsonResponse
from django.utils.text import capfirst

# from functools import update_wrapper
# from typing import Any
# from django.apps import apps

# from django.db.models.query import QuerySet
# from django.views.generic import ListView, CreateView, FormView, DetailView
from datetime import datetime
from django.contrib.admin import AdminSite
from django.contrib import admin 
# from django.contrib.admin.sites import site
from django.shortcuts import get_object_or_404, redirect, render
from .models import *
from django.db.models import Q
# from django.contrib.admin.options import ModelAdmin

# from django.contrib.admin.utils import flatten_fieldsets

# from django.contrib.admin.helpers import AdminForm, InlineAdminFormSet

# from django.contrib.admin.options import get_content_type_for_model

from django.contrib.auth.decorators import login_required

from tabel.models import Tabel
from .forms import *
from .decorators import allowed_users
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from tabel.models import TimeTrackingTabel

# Create your views here.

YEARS_CHOICES = (
    ('2023','2023'),
    ('2024','2024'),
    ('2025','2025')
)
def russian_month_to_int(month):
    month_names = {
        'Январь': 1,
        'Февраль': 2,
        'Март': 3,
        'Апрель': 4,
        'Май': 5,
        'Июнь': 6,
        'Июль': 7,
        'Август': 8,
        'Сентябрь': 9,
        'Октябрь': 10,
        'Ноябрь': 11,
        'Декабрь': 12
    }
    # Convert month to title case to handle different capitalizations
    month_title = month.strip().title()
    return month_names.get(month_title, None)
month_names_ru = {
    "January": "Январь",
    "February": "Февраль",
    "March": "Март",
    "April": "Апрель",
    "May": "Май",
    "June": "Июнь",
    "July": "Июль",
    "August": "Август",
    "September": "Сентябрь",
    "October": "Октябрь",
    "November": "Ноябрь",
    "December": "Декабрь"
}

#filter checking valid or not
def is_valid_queryparam(param):
    return param != '---' and param is not None

#show sidebar
def sidebar(request):
    admin_site = AdminSite()
    available_apps = []

    app_list = admin.site.get_app_list(request)
    for app in app_list:
        app_name = app['name']
        app_label = app['app_label']
        app_url = reverse('admin:index') + f"{app_label}/"
        models_list = []
        for model_dict in app['models']:
            model = model_dict.get('model')  # Get the model class if it exists
            model_admin = admin_site._registry.get(model)
            if model:
                app_label = model._meta.app_label
                info = (app_label, model._meta.model_name)
                model_info = {
                    "model": model,
                    "name": capfirst(model._meta.verbose_name_plural),
                    "object_name": model._meta.object_name,
                    # "perms": perms,
                    "admin_url": None,
                    "add_url": None,
                }
                model_info["admin_url"] = reverse(
                            "admin:%s_%s_changelist" % info, current_app=capfirst(model._meta.verbose_name_plural)
                        )
                model_info["add_url"] = reverse(
                            "admin:%s_%s_add" % info, current_app=capfirst(model._meta.verbose_name_plural)
                        )

                models_list.append(model_info)
        available_apps.append(
            {
                'name': app_name,
                'models': models_list,
                'app_label': app_label,
                'app_url': app_url
            }
        )

    data = {
        "has_permission": True,
        "is_popup": False,
        "available_apps": available_apps,
        "site_title": admin_site.site_title,
        "site_header": admin_site.site_header,
    }

    admin_context = admin_site.each_context(request)
    admin_context.update(data)
    return admin_context

def home(request):
    graph =  Graph.objects.all()
    reservoir_form = GraphReservoirForm()
    subdivision_form = GraphSubdivisionForm()
    years = YearSelectForm()

    #filters

    filter_year = request.GET.get('years')
    request.session['selected_year'] = filter_year

    filter_reservoir = request.GET.get('mesto')
    request.session['selected_reservoir'] = filter_reservoir

    filter_subdiv = request.GET.get('podrazd')
    request.session['selected_subdivision'] = filter_subdiv


    if is_valid_queryparam(filter_year):
        graph = graph.filter(year=filter_year)

    if is_valid_queryparam(filter_reservoir):
        graph = graph.filter(reservoir__name__icontains=filter_reservoir)

    if is_valid_queryparam(filter_subdiv):
        graph = graph.filter(subdivision__name__icontains=filter_subdiv)

        
    context = {
        'selected_year':request.session['selected_year'],
        "selected_reservoir": request.session['selected_reservoir'],
        "selected_subdivision": request.session['selected_subdivision'],        
        'graph':graph,
        'year_form':years,
        "reservoir":reservoir_form,
        "subdivision":subdivision_form,
    }
    return render(request,'graph/home.html',context)

@login_required(login_url='admin/login/')
@allowed_users(allowed_roles=['Табельщик', 'Администратор', 'Руководитель'])
def graph_admin(request):

    #chosen graph
    if 'graph_pk' in request.GET:
        graph_pk = request.GET['graph_pk']
        request.session['chosen_pk'] = graph_pk
        
    
    graph_pk = request.session['chosen_pk']  
    graph = Graph.objects.get(pk=graph_pk)
    json_base64 = None
    employees = graph.employees.all()
    # tracking = TimeTracking.objects.all().select_related('employee_id')
    month = graph.month
    year = graph.year
    tracking = TimeTracking.objects.filter(date__year=int(year), date__month=int(month), graph_id = graph).select_related('employee_id')
    name_month_en = calendar.month_name[int(month)]
    name_month_ru = month_names_ru[name_month_en]
    num_days = calendar.monthrange(int(year),int(month))[1]
    days = range(1,num_days+1)
    #attendace calculation start
    full_attendance = Attendance.objects.all().iterator()
    list_att = []
    for each in full_attendance:
        list_att.append(each.name)
    # print(list_att)
    directory = {}
    for employee in employees:
        pairs = [('worked_days', 0), ('weekends', 0), ('days_in_month', len(days)), ('total_work_hours', 0)]
        directory[int(f'{employee.tabel_number}')] = dict(pairs)

    for employee in employees:
        for work in tracking:
            if work.employee_id == employee:
                if str(work.worked_hours).isdigit():
                    directory[int(f'{work.employee_id.tabel_number}')]['worked_days'] += 1 
                    directory[int(f'{work.employee_id.tabel_number}')]['total_work_hours'] += int(work.worked_hours)
                    # directory[int(f'{employee.tabel_number}')]['worked_days'] += 1 
                    # directory[int(f'{employee.tabel_number}')]['total_work_hours'] += int(work.worked_hours)
                elif ('/' in work.worked_hours):
                    sep = work.worked_hours.split('/')
                    if len(sep)==2:
                        if str(sep[0]).isdigit():
                            directory[int(f'{work.employee_id.tabel_number}')]['total_work_hours'] += int(sep[0])
                            directory[int(f'{work.employee_id.tabel_number}')]['worked_days'] += 1 
                            if str(sep[1]).isdigit():
                                # directory[int(f'{work.employee_id.tabel_number}')]['night_work'] +=int(sep[1])
                                pass

                # elif not str(work.worked_hours).isdigit():
                else:
                    directory[int(f'{work.employee_id.tabel_number}')]['weekends'] += 1


    time_tracking_dict = {}
    for employee in employees:
        list = []
        for day in days:
            list.append(0)
        time_tracking_dict[int(f'{employee.tabel_number}')] = list

    for employee in employees:
        for work in tracking:
            if work.employee_id == employee:
                employee_id = work.employee_id.tabel_number
                day_index = days.index(work.date.day)
                time_tracking_dict[int(employee_id)][day_index] = work.worked_hours
    #creating table
    if request.method == 'POST':
        if 'approve_graph' in request.POST:
            pass
        if request.headers["content-type"].strip().startswith("application/json"):
            if 'subjectDn' in request.body.decode('utf-8'):
                data = json.loads(request.body)
                
                subject_dn = data.get('subjectDn')
                iin = None
                users_iin = request.user.iin
                # print("users: ",users_iin)
                if subject_dn:
                    match = re.search(r'IIN(\d{12})', subject_dn)
                    if match:
                        iin = match.group(1)
                        # print(iin)
                        if iin == users_iin:
                            # from django.contrib import messages
                            graph_pk = request.session['chosen_pk']
                            # print(graph_pk)
                            graph_inst = Graph.objects.get(pk=graph_pk)
                            print(graph_inst)
                            employees_graph = graph_inst.employees.all()
                            tracking = TimeTracking.objects.filter(date__year=int(graph_inst.year), date__month=int(graph_inst.month),graph_id=graph_inst).select_related('employee_id')
                            tabel_instance = None
                            # messages.success(request,'Chto-to')
                            try:
                                tabel_instance = Tabel.objects.get(
                                    reservoir=graph_inst.reservoir,
                                    subdivision=graph_inst.subdivision,
                                    month=graph_inst.month,
                                    year=graph_inst.year,
                                )
                                if tabel_instance:
                                    messages.error(request,'Табель уже согласован')
                            except ObjectDoesNotExist:
                                tabel_instance = Tabel.objects.create(
                                    reservoir=graph_inst.reservoir,
                                    subdivision=graph_inst.subdivision,
                                    month=graph_inst.month,
                                    year=graph_inst.year,
                                )
                                time_tracking_dict_tabel = {}
                                for employee in employees_graph:
                                    list = []
                                    for day in days:
                                        list.append(0)
                                    time_tracking_dict_tabel[int(f'{employee.tabel_number}')] = list
                                for work in tracking:
                                    employee_id = work.employee_id.tabel_number
                                    day_index = days.index(work.date.day)
                                    time_tracking_dict_tabel[int(employee_id)][day_index] = work.worked_hours

                                for employee in employees_graph:
                                    for work in tracking:
                                        if work.employee_id == employee:
                                            employee_id = work.employee_id.tabel_number
                                            day_index = days.index(work.date.day)
                                            time_tracking_dict_tabel[int(employee_id)][day_index] = work.worked_hours
                                    
                                time_tracking_tabel_instances = []
                                for employee_id, values in time_tracking_dict_tabel.items():
                                    for day_index, worked_hours in enumerate(values):
                                        time_tracking_tabel_instances.append(TimeTrackingTabel(
                                            employee_id=Employees.objects.get(tabel_number=employee_id),
                                            date=datetime.datetime(int(graph_inst.year), int(graph_inst.month), days[day_index]),
                                            worked_hours=worked_hours,
                                            tabel_id = tabel_instance
                                        ))
                                TimeTrackingTabel.objects.bulk_create(time_tracking_tabel_instances)
                                tabel_instance.employees.set(employees_graph)
                                # print(iin)
                                # graph_inst.status = 'Согласованный'
                                print(graph_inst.status)
                                graph_inst.save()
                                json_base64 = graph_to_json(graph_pk)
                                base64 = text_to_base64(json_base64)
                                messages.success(request,'Табель согласован')
                                return redirect('admin:tabel_tabel_changelist')
                        else:
                            messages.error(request,'Вы не можете согласовать график')
                            print('net')
                    # print("IIN: ",iin)
        
    #

    

    #
    context = {
        'graph_pk':graph_pk,
        "year":year,
        "month":month,
        'days':days,
        "selected_month": name_month_ru,
        'employees':employees,
        # 'attendance': attendance, 
        # 'no_attendance': no_attendance,
        'time_tracking': tracking,
        'graph':graph,
        'calculations': directory,
        'time_tracking_dict': time_tracking_dict,
        "json_base64":json_base64,
    }

    #adding admin side bar
    context.update(sidebar(request))
    return render(request,'graph/graph_admin.html',context)


@login_required(login_url='admin/login/')
@allowed_users(allowed_roles=['Табельщик', 'Администратор'])
def graph_admin_update(request):
    search_text = request.POST.get('time_tracking_set-prefix-employee_id')
    graph_pk = request.session['chosen_pk']
    graph = Graph.objects.get(pk=graph_pk)
    employees = graph.employees.all()
    # print("graph: ",graph)
    attendance_full = Attendance.objects.all().iterator()
    attendance = Attendance.objects.filter(type="дни явок")
    no_attendance = Attendance.objects.filter(type="дни неявок")
    
    # tracking = TimeTracking.objects.all().select_related('employee_id')
    month = int(graph.month)
    year = int(graph.year)
    tracking = TimeTracking.objects.filter(date__year=int(year), date__month=int(month), graph_id = graph).select_related('employee_id')

    # tracking = tracking.filter(date__year=int(year)).filter(date__month=int(month))
    employee_tabel_numbers = [employee.tabel_number for employee in employees]
    try:
        search_employee = Employees.objects.filter(
        Q(tabel_number__icontains=search_text) |
        Q(name__icontains=search_text)| 
        Q(surname__icontains=search_text) | 
        Q(middlename__icontains=search_text)
        )
    except:
        search_employee = Employees.objects.filter().exclude(
            tabel_number__in = employee_tabel_numbers
        )

    num_days = calendar.monthrange(int(year),int(month))[1]
    days = range(1,num_days+1)

    time_tracking_dict = {}
    for employee in employees:
        list = []
        for day in days:
            list.append(0)
        time_tracking_dict[int(f'{employee.tabel_number}')] = list
    # print(time_tracking_dict)

    for employee in employees:
        for work in tracking:
            if work.employee_id == employee:
                employee_id = work.employee_id.tabel_number
                day_index = days.index(work.date.day)
                time_tracking_dict[int(employee_id)][day_index] = work.worked_hours

    if request.method == 'POST':
        if request.headers["content-type"].strip().startswith("application/json"):
            if "employee_id_delete" in request.body.decode('utf-8'):
                employee_deletion_data = json.loads(request.body)
                employee_tabel_number = employee_deletion_data.get('employee_id_delete')
                employee_delete = Employees.objects.get(pk=employee_tabel_number)
                # employee_delete = employees.get(pk=employee_tabel_number)

                graph.employees.remove(employee_delete)
                # TimeTracking.objects.filter(employee_id = employee_tabel_number).delete()
                tracking.filter(employee_id = employee_tabel_number).delete()
                messages.success(request,f'Сотрудник {employee_delete.name} удален')

                    
            if "employee_id" in request.body.decode('utf-8'):
                employee_addition_data = json.loads(request.body)
                employee_tabel_number = employee_addition_data.get('employee_id')
                # employee_graph = employee_addition_data.get('graph_pk')
                employee_add = Employees.objects.get(pk=employee_tabel_number)
                for day in days:
                    # TimeTracking.objects.create(
                    #     employee_id = employee_add,
                    #     date = datetime.datetime(year, month, day),
                    #     worked_hours = ''
                    TimeTracking.objects.bulk_create(
                        [
                        TimeTracking(employee_id = employee_add,
                        date = datetime.datetime(year, month, day),
                        worked_hours = '',
                        graph_id = graph
                        )
                        ]
                    )
                graph.employees.add(employee_tabel_number)
                messages.success(request,f'Сотрудник {employee_add.name} добавлен')
    
        for key,value in request.POST.items():
            if key.startswith('worked_hours_'):
                time_tracking_day = int(key.split('_')[4])
                key_parts = key.split('_')
                if len(key_parts) >=5:
                    for day in days:
                        time_tracking_employee = int(key.split('_')[2])
                        if day == time_tracking_day:
                            day_index = days.index(time_tracking_day)
                            time_tracking_dict[time_tracking_employee][day_index] = value
                        else:
                            if time_tracking_employee not in time_tracking_dict:
                                time_tracking_dict[time_tracking_employee] = [0] * len(days)
                            
                            # if time_tracking_day in days:
                            #     day_index = days.index(time_tracking_day)
                            #     time_tracking_dict[time_tracking_employee][day_index] = value
                else:
                    day_index = days.index(time_tracking_day)
                    time_tracking_dict[time_tracking_employee][day_index] = value
        #saving tracking 
        for work in tracking:
            employee_id = work.employee_id.tabel_number
            if employee_id in time_tracking_dict:
                # print(f'{employee_id} is in time_tracking_dict')
                day = work.date.day
                if day in days:
                    day_index = days.index(day)
                    w = time_tracking_dict[employee_id][day_index]
                    if w is not None:
                        if w == '':
                            work.worked_hours = 0 #
                        else:
                            work.worked_hours = w
        # Batch save all the updated works
        TimeTracking.objects.bulk_update(tracking, ['worked_hours'])


        # print(time_tracking_dict)
        # print(time_tracking_dict)
        return redirect(reverse('graph:graph_admin') +f'?graph_pk={graph_pk}')
    
    name_month_en = calendar.month_name[int(month)]
    name_month_ru = month_names_ru[name_month_en]
    

    #attendance calculation
    directory = {}
    for employee in employees:
        pairs = [('worked_days', 0), ('weekends', 0), ('days_in_month', len(days)), ('total_work_hours', 0)]
        directory[int(f'{employee.tabel_number}')] = dict(pairs)

    for employee in employees:
        for work in tracking:
            if work.employee_id == employee:
                if str(work.worked_hours).isdigit():
                    directory[int(f'{work.employee_id.tabel_number}')]['worked_days'] += 1 
                    directory[int(f'{work.employee_id.tabel_number}')]['total_work_hours'] += int(work.worked_hours)
                    # directory[int(f'{employee.tabel_number}')]['worked_days'] += 1 
                    # directory[int(f'{employee.tabel_number}')]['total_work_hours'] += int(work.worked_hours)
                elif ('/' in work.worked_hours):
                    sep = work.worked_hours.split('/')
                    if len(sep)==2:
                        if str(sep[0]).isdigit():
                            directory[int(f'{work.employee_id.tabel_number}')]['total_work_hours'] += int(sep[0])
                            directory[int(f'{work.employee_id.tabel_number}')]['worked_days'] += 1 
                            if str(sep[1]).isdigit():
                                # directory[int(f'{work.employee_id.tabel_number}')]['night_work'] +=int(sep[1])
                                pass

                # elif not str(work.worked_hours).isdigit():
                else:
                    directory[int(f'{work.employee_id.tabel_number}')]['weekends'] += 1
                    # directory[int(f'{employee.tabel_number}')]['weekends'] += 1
    context = {
        'employees_all': search_employee,
        'graph_pk':graph_pk,
        "year":year,
        "month":month,
        'days':days,
        "selected_month": name_month_ru,
        'employees':employees,
        'attendance': attendance,
        'attendance_full':attendance_full, 
        'no_attendance': no_attendance,
        'time_tracking': tracking,
        'graph':graph,
        'calculations': directory,
        'time_tracking_dict': time_tracking_dict,

    }
    
    #adding admin side bar
    context.update(sidebar(request))
    # graph_update_sidebar = {
    #     "is_nav_sidebar_enabled": False,
    #     "is_popup": True,
    #     }
    # context.update(graph_update_sidebar)

    return render(request,'graph/graph_admin_update.html',context)

from openpyxl import load_workbook as lw

def upload_file(request):
    table_data = {
        'month': '',
        'data': [],
        'formatted': False
    }
    months = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь', 'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь']
    if request.method == 'POST' and request.FILES['file']:
        uploaded_file = request.FILES['file']
        # Do something with the uploaded file
        if uploaded_file.name.endswith('.xlsx'):
            load_workbook = lw(uploaded_file,data_only=True)
            graph = load_workbook.active

            data_length = 0
            while isinstance(graph['A' + str(5 + data_length)].value, int):
                data_length += 1
            # print(data_length)
            row_num = 4
            list_values = [str(x.value).lower() for x in graph['A']]
            list_values = list_values[3:]
            month = 1
            for j in range(1,len(months)+1):
                for value in list_values:
                    if any(month in value for month in months):
                        table_data['month'] = value.capitalize()
                        list_values = list_values[data_length+1:]
                        break
                for n in range(data_length+1):
                    row_num +=1
                    list = [x.value for x in graph[row_num]]
                    if list.count(None) > 3:
                        break
                    table_data['data'].append(list) 
                for i in table_data['data']:
                    #adding job start
                    job_name = i[2].split(' ', 1)
                    if len(job_name) != 2:
                        job_name.append(' ')
                    elif job_name[0] == '':
                        job_name_temp = job_name[0].split(' ',1)
                        if len(job_name_temp) == 2:
                            job_name[0]=job_name_temp[0]
                            job_name[1]=job_name_temp[1]
                        else:
                            job_name[0]=job_name_temp[0]
                            job_name[1]=' '
                    try:
                        job_instance = Job.objects.get(name=job_name[0])
                    except:
                        if job_name[0] != '':  
                            Job.objects.create(
                                name=job_name[0],
                                description=job_name[1]
                            )
                #adding job end

                    ###
                #adding oil_place start
                for i in table_data['data']:
                    
                    reservoir_name = i[3]
                    try:
                        reservoir_inst = OilPlace.objects.get(name=reservoir_name)
                    except:
                        OilPlace.objects.create(
                            name=reservoir_name
                        )

                #adding oil_place end
                
                #adding employee start
                count=1
                for i in table_data['data']:
                    
                    employee_excel = i[1].split()
                    if len(employee_excel) == 2:
                        employee_excel.append(' .')
                    elif len(employee_excel) == 1:
                        employee_excel.append(' .')
                        employee_excel.append(' .')
                    try:
                        employee_inst = Employees.objects.get(
                            name = employee_excel[0],
                            surname = employee_excel[1],
                            middlename = employee_excel[2]
                        )
                    except:
                        empl_job = i[2].split(' ',1)
                        if empl_job:
                            try:
                                job_inst = Job.objects.get(
                                    name=empl_job[0]
                                )
                            except:
                                continue
                        empl_oilplace = i[3]
                        if empl_oilplace:
                            try:
                                oilplace_inst = OilPlace.objects.get(
                                    name=empl_oilplace
                                )
                            except:
                                continue
                        tarrif_category = len(i[4])

                        Employees.objects.create(
                            tabel_number = count,
                            name = employee_excel[0],
                            surname = employee_excel[1],
                            middlename = employee_excel[2],
                            tariff_category=tarrif_category,
                            job=job_inst,
                            oil_place=oilplace_inst
                        )
                        count+=1
                    #employee add end                    
                table_data_month = table_data['month'].split()
                # print(table_data_month)
                # print(len(table_data_month))
                graph_month = russian_month_to_int(table_data_month[0])
                graph_year =  int(table_data_month[1])                    
                graph_inst = None
                employee_inst = None
                try:
                    graph_inst = Graph.objects.get(
                        month= graph_month,
                        year = graph_year
                    )
                except:
                    oilplace_inst = OilPlace.objects.get(name='Ботахан')
                    subdivision_inst = Subdivision.objects.get(name='БДН')
                    Graph.objects.create(
                        month = graph_month,
                        year = graph_year,
                        reservoir = oilplace_inst,
                        subdivision = subdivision_inst
                    )
                for row in table_data['data']:
                    employee_excel = row[1].split()
                    if len(employee_excel) == 2:
                        employee_excel.append(' .')
                    elif len(employee_excel) == 1:
                        employee_excel.append(' .')
                        employee_excel.append(' .')
                    try:
                        employee_inst = Employees.objects.get(
                            name = employee_excel[0],
                            surname = employee_excel[1],
                            middlename = employee_excel[2]
                        )
                    except:
                        continue
                    if graph_inst is not None and not graph_inst.employees.filter(tabel_number=employee_inst.tabel_number).exists():
                        graph_inst.employees.add(employee_inst)
                    graph_inst.save()

                    table_data_month = table_data['month'].split(' ')
                    #timetracking add start
                    for row in table_data['data']:
                        count_day = 1
                        length = len(row)
                        employee_excel = row[1].split()
                        if len(employee_excel) == 2:
                            employee_excel.append(' .')
                        elif len(employee_excel) == 1:
                            employee_excel.append(' .')
                            employee_excel.append(' .')
                            try:
                                employee_inst = Employees.objects.get(
                                    name = employee_excel[0],
                                    surname = employee_excel[1],
                                    middlename = employee_excel[2]
                                )
                            except:
                                continue
                        graph_year = 2023
                        if len(table_data_month) == 3:
                            graph_year = int(table_data_month[1])
                        else:
                            graph_year = int(table_data_month[2])
                        graph_month = russian_month_to_int(table_data_month[0])
                        
                        # print('graph_month ',graph_month)
                        # Get the number of days in the current month
                        num_days_in_month = calendar.monthrange(graph_year, graph_month)[1]
                        
                        for value in row[5:length-4]:
                            # while count_day <= num_days_in_month:
                                try:
                                    timetracking_inst = TimeTracking.objects.get(
                                        employee_id = employee_inst,
                                        date=datetime.datetime(2023,graph_month,count_day),
                                        graph_id = graph_inst
                                    )
                                except:
                                    # print('month ',month,' day ',count_day)
                                    if value is not None:
                                        TimeTracking.objects.create(
                                            employee_id = employee_inst,
                                            date=datetime.datetime(2023,graph_month,count_day),
                                            worked_hours = value,
                                            graph_id = graph_inst
                                        )
                                count_day+=1
                            
                            # if  count_day > num_days_in_month:
                            #     count_day=1

                    graph_month = (graph_month % 12) + 1
                # print(table_data)
                table_data['data'].clear()
        else:
            messages.error(request,'файл не xlsx формата :(')
        # messages.error(request,'Вы не прикрепили файл')
    context = {}
    context.update(sidebar(request))
    return render(request,'graph/parsing_graph.html',context)

def graph_to_json(graph_pk):
    """
    "graph":{
        graph_pk:{
            "year":year,
            "month":month,
            "reservoir":reservoir,
            "subdivision":subivision,
            "status":status,
            "employees":{
                employee_tabel_number:{
                    "name":name,
                    "surname":surname,
                    "middlename":middlename,
                    "tariff_category":tariff_category,
                    "job":job,
                    "oil_place":oil_place,
                    timetracking : {
                        date:worked_hours,
                        },
                    "total_days_worked":total_days_worked,
                    "total_rest_days":total_rest_days,
                    "total_days_in_month":total_days_in_month",
                    "total_hours":total_hours
                }
            }
        }
        
    }
    """
    graph = Graph.objects.get(pk=graph_pk)
    employees = graph.employees.all()
    reservoir = graph.reservoir.name
    subdivision = graph.subdivision.name
    month = graph.month
    year = int(graph.year)
    num_days = calendar.monthrange(int(year),int(month))[1]
    days = range(1,num_days+1)
    tracking = TimeTracking.objects.filter(date__year=year, date__month=int(month), graph_id = graph).select_related('employee_id')
    graph_json = {
    "graph": {
        str(graph_pk): {
            "year": year,
            "month": month,
            "reservoir": f"{reservoir}",
            "subdivision": f"{subdivision}",
            "status": graph.status,
            "employees": {}
            }
        }
    }
    time_tracking_dict = defaultdict(lambda: {f"{day}.{month}.{year}": 0 for day in days})

    for work in tracking:
        employee_id = work.employee_id.tabel_number
        day_str = f"{work.date.day}.{month}.{year}"
        time_tracking_dict[employee_id][day_str] = work.worked_hours
    
    directory = {}
    for employee in employees:
        pairs = [('worked_days', 0), ('weekends', 0), ('days_in_month', len(days)), ('total_work_hours', 0)]
        directory[int(f'{employee.tabel_number}')] = dict(pairs)

    for employee in employees:
        for work in tracking:
            if work.employee_id == employee:
                if str(work.worked_hours).isdigit():
                    directory[int(f'{work.employee_id.tabel_number}')]['worked_days'] += 1 
                    directory[int(f'{work.employee_id.tabel_number}')]['total_work_hours'] += int(work.worked_hours)
                    # directory[int(f'{employee.tabel_number}')]['worked_days'] += 1 
                    # directory[int(f'{employee.tabel_number}')]['total_work_hours'] += int(work.worked_hours)
                elif ('/' in work.worked_hours):
                    sep = work.worked_hours.split('/')
                    if len(sep)==2:
                        if str(sep[0]).isdigit():
                            directory[int(f'{work.employee_id.tabel_number}')]['total_work_hours'] += int(sep[0])
                            directory[int(f'{work.employee_id.tabel_number}')]['worked_days'] += 1 
                            if str(sep[1]).isdigit():
                                # directory[int(f'{work.employee_id.tabel_number}')]['night_work'] +=int(sep[1])
                                pass

                # elif not str(work.worked_hours).isdigit():
                else:
                    directory[int(f'{work.employee_id.tabel_number}')]['weekends'] += 1
    
    # print(time_tracking_dict)
    for employee in employees:
        job = Job.objects.get(pk=employee.job.pk)
        
        employee_dict = {
            "name": employee.name,
            "surname": employee.surname,
            "middlename": employee.middlename,
            "tariff_category": employee.tariff_category,
            "job": employee.job.name,  
            "oil_place": employee.oil_place.name,  
            "time_tracking": time_tracking_dict[employee.tabel_number],
            "worked_days":directory[employee.tabel_number]['worked_days'],
            "weekends":directory[employee.tabel_number]['weekends'],
            "total_days_in_month":len(days),
            "total_work_hours":directory[employee.tabel_number]['total_work_hours'],
            
        }
        graph_json["graph"][str(graph_pk)]["employees"][employee.tabel_number] = employee_dict
    path = f'../tabulation/graph_json.json'
    with open(path,'w',encoding='utf-8') as json_file:
        json.dump(graph_json,json_file,ensure_ascii=False,indent=4)
    # print(json.dumps(graph_json,ensure_ascii=False,indent=4))
    json_str = json.dumps(graph_json,ensure_ascii=False,indent=4)
    import base64
    json_base64 = base64.b64encode(json_str.encode('utf-8')).decode('utf-8')
    return json_base64

def text_to_base64(text):
    import base64
    text_base64 = base64.b64encode(text.encode('utf-8')).decode('utf-8')
    return text_base64




